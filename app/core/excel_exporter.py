"""Export Excel tổng hợp toàn bộ hồ sơ trong 1 lần ghi (tối ưu I/O).

Kế thừa `ghi_excel_mot_lan_toi_uu` cũ nhưng cột lấy từ `style.row_mapping`
(bỏ hardcode 7 cột). Mỗi hồ sơ: 2 dòng tiêu đề + header bảng + dữ liệu.
"""

from __future__ import annotations

from typing import Any, Dict, List

from app.models.style import StyleConfig

_FONT = "Times New Roman"

# Ký tự mở đầu ô có thể bị Excel diễn giải thành công thức (CSV/formula injection).
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _is_number(s: str) -> bool:
    """True nếu chuỗi là số hợp lệ (để không bọc nhầm số âm '-5')."""
    try:
        float(s)
        return True
    except ValueError:
        return False


def _sanitize_cell(value: Any) -> Any:
    """Chống formula injection: ô chuỗi bắt đầu bằng `= + - @` → thêm `'` đứng trước.

    Bỏ qua nếu chuỗi thực chất là một con số (vd `-5`) để không làm hỏng giá trị số.
    """
    if isinstance(value, str) and value[:1] in _FORMULA_PREFIXES and not _is_number(value):
        return "'" + value
    return value


def export_excel(all_groups: List[Dict[str, Any]], out_path, style: StyleConfig) -> bool:
    """Ghi tất cả nhóm ra 1 file Excel.

    Args:
        all_groups: list `{"ho_so_so": str, "records": [dict,...]}` theo thứ tự.
        out_path: đường dẫn file .xlsx.
        style: để lấy header cột (`row_mapping`) và tiêu đề (`settings.tieu_de`).
    Returns: True nếu ghi thành công, False nếu không có dữ liệu.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    if not all_groups:
        return False

    # Cột: nhãn hiển thị = tên cột sheet (value của row_mapping), theo thứ tự columns.
    headers = _excel_headers(style)
    columns_sheet = list(style.row_mapping.values())
    tieu_de = str(style.settings.get("tieu_de", "MỤC LỤC VĂN BẢN, TÀI LIỆU"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mục Lục Hồ Sơ"

    row_num = 1
    for item in all_groups:
        ho_so_so = item["ho_so_so"]
        records = item["records"]

        c = ws.cell(row=row_num, column=4, value=tieu_de)
        c.font = Font(name=_FONT, size=13, bold=True)
        c.alignment = Alignment(horizontal="center")
        c.number_format = "@"
        row_num += 1

        c = ws.cell(
            row=row_num,
            column=4,
            value=f"Số, ký hiệu hồ sơ (đơn vị bảo quản): {ho_so_so}",
        )
        c.font = Font(name=_FONT, size=13, bold=True, color="FF0000")
        c.alignment = Alignment(horizontal="center")
        c.number_format = "@"
        row_num += 1

        for col_num, header in enumerate(headers, start=1):
            c = ws.cell(row=row_num, column=col_num, value=header)
            c.font = Font(name=_FONT, size=10, bold=True)
            c.alignment = Alignment(horizontal="center")
            c.number_format = "@"
        row_num += 1

        for rec in records:
            for col_num, col_name in enumerate(columns_sheet, start=1):
                val = _sanitize_cell(rec.get(col_name, ""))
                c = ws.cell(row=row_num, column=col_num, value=val)
                c.font = Font(name=_FONT, size=10)
                c.number_format = "@"
                # Cột "trích yếu" canh trái (nếu có), còn lại canh giữa.
                is_trich_yeu = _is_trich_yeu_column(style, col_num)
                c.alignment = Alignment(
                    horizontal="left" if is_trich_yeu else "center"
                )
            row_num += 1

        row_num += 1  # dòng trống ngăn cách các hồ sơ

    wb.save(str(out_path))
    return True


def _excel_headers(style: StyleConfig) -> List[str]:
    """Nhãn header Excel = tên cột sheet theo thứ tự `row_mapping`."""
    return list(style.row_mapping.values())


def _is_trich_yeu_column(style: StyleConfig, col_num_1based: int) -> bool:
    """Cột thứ `col_num` có phải biến 'trich_yeu' không (để canh trái)."""
    variables = list(style.row_mapping.keys())
    idx = col_num_1based - 1
    return 0 <= idx < len(variables) and variables[idx] == "trich_yeu"
